def cerca(file_in):
    tabfinale=[]
    f=open(file_in,"r",encoding="UTF-8")
    frase=f.read().lower()
    ind=frase.find("\"contract\":\"sale\",")
    vistamappa=0
    if ind<=0:
        ind=frase.find("in-realestateresults__item\" id")
        frase=frase[ind:]
        lista_parole=frase.split("in-realestateresults__item\" id")
        vistamappa=1
    else:
        frase=frase[ind:]
        #print(frase)
        lista_parole=frase.split("\"contract\":\"sale\",")
    conta=0
    maxc=0
    res=[]

    for parola in lista_parole:
        if vistamappa==0:
            if len(parola)>10:
                ival=parola.find("\"floor\":{\"abbreviation\":")
                if ival>0:
                    parola=parola[ival+24:]
                    piano=parola[0:parola.find(",\"value\":")]
                    piano=piano.replace("\"","")
                    pianodesc=parola[parola.find(",\"value\":")+10:parola.find("\"},\"multimedia")]
                ival=parola.find("\"price\":")
                parola=parola[ival:]
                prezzo=parola[parola.find("\"value\":")+8:parola.find(",\"formattedvalue\"")]
                prezzo=prezzo.removeprefix("\"value\":")
                ival=parola.find("\"rooms\":")
                locali=parola[ival+9:parola.find("\",\"shop\":")]
                ival=parola.find("\"surface\":")
                parola=parola[ival:]
                mq=parola[0:parola.find(",\"surfacevalue\"")]
                mq=mq.removeprefix("\"surface\":\"")
                r=mq.find(" m")
                mq=mq[0:r]
                try:
                    p=int(prezzo)
                    q=int(mq)   
                    prmq=p/q
                    prmq=round(prmq,2)
                except:
                    prmq=0
                ##########locali-superficie-piano-anno costr-nome ag.
                ##########"rooms":"5+"
                ##########"surface":"503 m²"
                ##########"floor":{"abbreviation":"2","value":"2 piani: 2°, con ascensore"               
                #"agency":{"id":48026,"type":"agency","label":"agenzia","displayName":"CASANOVA IMMOBILIARE"
                res.append((prezzo, mq, prmq, locali, piano, pianodesc))
        else:
            l=len(parola)
            if l>75:
                ival=parola.find("main in-realestatelistcard__features--main")
                parola=parola[ival:]
                prezzo=parola[parola.find(">")+1:parola.find("</li>")]
                prezzo=prezzo.removeprefix("€ ")
                prezzo=prezzo.replace(".","")
                ival=parola.find("#size")
                parola=parola[ival:]
                mq=parola[19:parola.find("<span")]
                try:
                    p=int(prezzo)
                    q=int(mq)   
                    prmq=p/q
                    prmq=round(prmq,2)
                except:
                    prmq=0
                res.append((prezzo, mq, prmq))
                ####res.append((prezzo, mq, prmq, locali, piano, pianodesc))
                #planimetry"></use></svg>3</div>
                #stairs"></use></svg>T</div>
                #in-realEstateListCard__referent--image"><img class="nd-figure__content nd-ratio__img" loading="lazy" src="./lista_app2_files/1115582000.jpg" alt="TEMPOCASA SAN BENEDETTO DEL TRONTO – PORTO D’ASCOLI" 


                
    res.sort()
    if file_in=="lista_app.html":
        sheetimm['A1'] = "Prezzo"
        sheetimm['B1'] = "MQ"    
        sheetimm['C1'] = "Prezzo/MQ"    
        sheetimm['D1'] = "Vani"    
        sheetimm['E1'] = "Piano"    
        sheetimm['F1'] = "Piano descr."    
    for t in res:
        sheetimm.append(t)
        #print(t[0] + " - " + t[1] + " - " + str(t[2]))
    f.close()
    return res

def categorieOMI(file_omi):
    tabfinale=[]
    with open(file_omi,"r",encoding="iso8859-15") as f:
        tabella=0
        inizio=0
        puoistampare=0
        geopoi=0
        for line in f:
            if line.find("Geopoi")>0:
                geopoi=1
            ind=0
            ind=line.find("Risultato interrogazione:")
            if ind>0:
                ind2=line.find("</h4")
                periodo=line[ind+35:ind2]
                if geopoi==0:
                    periodo=periodo.replace("<br>","")               
            ind=line.find("Provincia:")
            if ind>0:
                if geopoi==0:
                    ind2=line.find("</span></p>")
                    prov=line[ind+40:ind2]
                else:
                    ind2=line.find("</span></p>")
                    prov=line[ind+37:ind2]
            ind=line.find("Comune:")
            if ind>0:
                if geopoi==0:
                    ind2=line.find("</span></p>")
                    comune=line[ind+37:ind2]
                else:
                    line=line[ind:]
                    ind2=line.find("</span></p>")
                    comune=line[34:ind2]
            ind=line.find("Fascia/zona:")
            if ind>0:
                if geopoi==0:
                    ind2=line.find("</span></p>")
                    fascia=line[53:ind2]
                else:
                    line=line[ind:]
                    ind2=line.find("</div><p>")
                    fascia=line[105:ind2]
            ind=line.find("Codice")
            if ind>0:
                if line.find("zona")>0:
                    if geopoi==0:
                        ind2=line.find("</span></p>")
                        zona=line[56:ind2]
                    else:
                        line=line[ind:]
                        ind2=line.find("</span></p>")
                        zona=line[39:ind2]
            if tabella==0:
                inizio=line.find("summary")
                if inizio>0:
                    tabella=1
                    #finetabella=0
            else:    
                #print(line)
                test=line.find("Tipologia")
                if geopoi==0:
                    ind=line.find("<tbody><tr><td class=\"sin\">")
                    if ind>0:
                        puoistampare=1
                    indend=line.find("</td><td class=\"sin\">")
                    tipologia=line[ind+27:indend]
                    ind=indend
                    indend=line.find("</td><td class=\"dx\" headers=\"vm vmmin\">")
                    stato=line[ind+21:indend]
                    ind=indend
                    indend=line.find("</td><td class=\"dx\" headers=\"vm vmmax\">")
                    valmin=line[ind+39:indend]
                    ind=indend
                    indend=line.find("</td><td class=\"center\">")
                    valmax=line[ind+39:indend]
                    ind=line.find("</td><td class=\"dx\" headers=\"vl vlmax\">")
                    indend=line.find("</td><td class=\"dx\" headers=\"vl vlmax\">",ind+10)
                    vallocmin=line[ind+39:indend]
                    ind=line.find("</td><td class=\"dx\" headers=\"vl vlmax\">",ind+10)
                    indend=line.find("</td><td class=\"center\">", ind+10)
                    vallocmax=line[ind+39:indend]
                    lista=[tipologia, stato, valmin, valmax, vallocmin, vallocmax]
                    if puoistampare==1:
                        tabfinale.append(lista)
                        #print(lista)
                else:
                    ind=line.find("<tr><td id=\"sin\">")
                    if ind>0:
                        puoistampare=1
                    while ind>0:
                        indend=line.find("</td><td id=\"sin\">")
                        tipologia=line[ind+17:indend].replace("&nbsp;","")
                        ind=indend
                        indend=line.find("</td><td id=\"dx\" headers=\"vm vmmin\">")
                        stato=line[ind+18:indend].replace("&nbsp;","")
                        ind=indend
                        indend=line.find("</td><td id=\"dx\" headers=\"vm vmmax\">")
                        valmin=line[ind+36:indend].replace("&nbsp;","")
                        ind=indend
                        indend=line.find("</td><td id=\"center\">")
#</td><td id="center">L&nbsp;</td></tr><tr><td id="sin">Abitazioni civili&nbsp;</td><td id="sin">Normale&nbsp;</td><td id="dx" headers="vm vmmin">1300&nbsp;</td><td id="dx" headers="vm vmmax">1900&nbsp;</td><td id="center">L&nbsp;</td><td id="dx" headers="vl vlmin">4&nbsp;</td><td id="dx" headers="vl vlmax">6&nbsp;</td><td id="center">L&nbsp;</td></tr><tr><td id="sin">Abitazioni di tipo economico&nbsp;</td><td id="sin">Normale&nbsp;</td><td id="dx" headers="vm vmmin">1150&nbsp;</td><td id="dx" headers="vm vmmax">1650&nbsp;</td><td id="center">L&nbsp;</td><td id="dx" headers="vl vlmin">3,5&nbsp;</td><td id="dx" headers="vl vlmax">5,2&nbsp;</td><td id="center">L&nbsp;</td></tr><tr><td id="sin">Box&nbsp;</td><td id="sin">Normale&nbsp;</td><td id="dx" headers="vm vmmin">1700&nbsp;</td><td id="dx" headers="vm vmmax">2550&nbsp;</td><td id="center">L&nbsp;</td><td id="dx" headers="vl vlmin">5,4&nbsp;</td><td id="dx" headers="vl vlmax">7,7&nbsp;</td><td id="center">L&nbsp;</td></tr><tr><td id="sin">Posti auto coperti&nbsp;</td><td id="sin">Normale&nbsp;</td><td id="dx" headers="vm vmmin">1150&nbsp;</td><td id="dx" headers="vm vmmax">1700&nbsp;</td><td id="center">L&nbsp;</td><td id="dx" headers="vl vlmin">3,4&nbsp;</td><td id="dx" headers="vl vlmax">5,1&nbsp;</td><td id="center">L&nbsp;</td></tr><tr><td id="sin">Posti auto scoperti&nbsp;</td><td id="sin">Normale&nbsp;</td><td id="dx" headers="vm vmmin">700&nbsp;</td><td id="dx" headers="vm vmmax">1050&nbsp;</td><td id="center">L&nbsp;</td><td id="dx" headers="vl vlmin">2,3&nbsp;</td><td id="dx" headers="vl vlmax">3,4&nbsp;</td><td id="center">L&nbsp;</td></tr><tr><td id="sin">Ville e Villini&nbsp;</td><td id="sin">Normale&nbsp;</td><td id="dx" headers="vm vmmin">1550&nbsp;</td><td id="dx" headers="vm vmmax">2200&nbsp;</td><td id="center">L&nbsp;</td><td id="dx" headers="vl vlmin">5,3&nbsp;</td><td id="dx" headers="vl vlmax">7,2&nbsp;</td><td id="center">L&nbsp;</td></tr>\t\t\t<tr>\n
                        valmax=line[ind+36:indend].replace("&nbsp;","")
                        ind=line.find("</td><td id=\"dx\" headers=\"vl vlmin\">")
                        indend=line.find("</td><td id=\"dx\" headers=\"vl vlmax\">",ind+10)
                        vallocmin=line[ind+36:indend].replace("&nbsp;","")
                        ind=line.find("</td><td id=\"dx\" headers=\"vl vlmax\">",ind+10)
                        indend=line.find("</td><td id=\"center\">", ind+10)
                        vallocmax=line[ind+36:indend].replace("&nbsp;","")
                        lista=[tipologia, stato, valmin, valmax, vallocmin, vallocmax]
                        tabfinale.append(lista)
                        #print(lista)
                        line=line[indend:]
                        ind=line.find("<tr><td id=\"sin\">")
                ind=line.find("</table>")
                if ind>=0:
                    #finetabella=1
                    tabella=0                                   
    sheet['A1'] = periodo
    sheet['B1'] = "Provincia"
    sheet['B2'] = "Fascia"
    sheet['B3'] = "Comune"
    sheet['B4'] = "Zona"
    sheet['C1'] = prov
    sheet['C2'] = fascia
    sheet['C3'] = comune
    sheet['C4'] = zona
    sheet.append(["Tipologia","Stato","Pr. min", "Pr. max", "Pr.Aff. min","Pr.Aff. max"])
    for elem in tabfinale:
        sheet.append(elem)
    nuovo_file.save("Immobiliare.xlsx")
    f.close()

import time 
import openpyxl
import os
import sys
"""
print("*******************MANUALE D'USO**************************************")
print("* SALVARE LE PAGINE DI RICERCA SU IMMOBILIARE.IT COME LISTA_APP.HTML *")
print("* SE NE SONO PIU DI UNA CHIAMARE LE ALTRE LISTA_APP(2, 3, 4).HTML    *")
print("* SALVARE LE PAGINE DI RICERCA SU OMI DI AG. ENTRATE COME OMI.HTML   *")
print("* ELIMINARE PRIMA LE VECCHIE PAGINE SCARICATE MAGARI SALVANDOLE      *")
print("*******************MANUALE D'USO**************************************")

esegui=input("Vuoi eseguire il calcolo?(S/N)")
"""
#script, arg = sys.argv

esegui="S"
os.chdir("immobilix")
if esegui.upper()=="S":
    #print("Elaborazione in corso...")     
    nuovo_file = openpyxl.Workbook()
    nuovo_file.save("Immobiliare.xlsx")
    sheetimm = nuovo_file.active
    if os.path.exists('./lista_app.html'):
        sheetimm.title="Dati Immobiliare" # = nuovo_file.create_sheet("Dati Immobiliare")
        cerca("lista_app.html")
    nuovo_file.save("Immobiliare.xlsx")
    if os.path.exists('lista_app2.html'):
        sheetimm = nuovo_file.active #nuovo_file.create_sheet("Dati Immobiliare 2")
        cerca("lista_app2.html")
    nuovo_file.save("Immobiliare.xlsx")
    if os.path.exists('lista_app3.html'):
        sheetimm = nuovo_file.active #nuovo_file.create_sheet("Dati Immobiliare 2")
        cerca("lista_app3.html")
    nuovo_file.save("Immobiliare.xlsx")
    if os.path.exists('lista_app4.html'):
        sheetimm = nuovo_file.active #nuovo_file.create_sheet("Dati Immobiliare 2")
        cerca("lista_app4.html")
    nuovo_file.save("Immobiliare.xlsx")
    if os.path.exists('OMI.html'):
        sheet=nuovo_file.create_sheet("Dati OMI")
        categorieOMI("OMI.html")
    nuovo_file.save("Immobiliare.xlsx")
    time.sleep(1.5)
    print("Elaborazione terminata...")     
    

